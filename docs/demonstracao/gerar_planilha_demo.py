# -*- coding: utf-8 -*-
"""
MOBGOV — gerador da planilha de DEMONSTRAÇÃO da secretaria.

Diferente de `dados/planilha_exemplo.py`, que é dependência zero e roda nos
testes, este script usa openpyxl para produzir um arquivo com a cara de um
arquivo de prefeitura de verdade: título mesclado, cabeçalho formatado,
colunas dimensionadas, painel congelado e um total com fórmula no rodapé.

Serve para a demonstração ao vivo — o gestor abre no Excel dele, reconhece o
formato, e aí vê o importador engolir o arquivo com a bagunça toda.

É ferramenta de apoio, não faz parte do sistema: exige `pip install openpyxl`.

Uso:
    python docs/demonstracao/gerar_planilha_demo.py
"""
from __future__ import annotations

import os
import random
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "..", ".."))

from dados.municipio_modelo import DISTRITOS, ESCOLAS, TIPOS_VEICULO

SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "planilha-prefeitura-demo.xlsx")
ALUNOS = 300
rng = random.Random(2026)

CABECALHO = ["Matrícula", "ALUNO(A)", "Endereço", "Nº", "Bairro", "Escola",
             "Turno", "Cadeira de Rodas", "Acompanhante", "Latitude",
             "Longitude", "Observações"]

NOMES = ["Ana", "Bruno", "Carla", "Diego", "Elis", "Fábio", "Gabriela",
         "Heitor", "Ivone", "João", "Karina", "Lucas", "Marina", "Nelson",
         "Olívia", "Paulo", "Queila", "Rafael", "Sônia", "Tiago", "Úrsula",
         "Vinícius", "Wesley", "Yasmin"]
SOBRENOMES = ["Silva", "Souza", "Oliveira", "Pereira", "Costa", "Almeida",
              "Rodrigues", "Nascimento", "Araújo", "Ferreira", "Gomes",
              "Barbosa", "Ribeiro", "Martins"]
RUAS = ["Rua das Acácias", "Estrada do Assentamento", "Av. Central",
        "Rua Projetada A", "Estrada Municipal RM-040", "Travessa São José",
        "Sítio Boa Vista", "Rua do Córrego", "Linha 3 do Assentamento"]
TURNOS = ["Manhã", "MAT", "manha", "M", "Tarde", "T", "Vespertino", "tarde "]
CADEIRA = ["", "", "", "", "-", "não", "NÃO", "N", "x", "SIM", "?"]
PESOS_CADEIRA = [34, 26, 14, 10, 5, 4, 3, 2, 1, 1, 1]
OBSERVACOES = ["", "", "", "", "porteira fechada, buzinar",
               "estrada ruim em dia de chuva", "mora depois da ponte",
               "descer no ponto do vizinho", "família mudou de endereço"]

TINTA = "FF1F3864"
CINZA = "FFF2F2F2"


def _linhas_de_alunos() -> list:
    distritos = list(DISTRITOS.items())
    linhas = []
    for i in range(ALUNOS):
        bairro, (clat, clon, raio, *_) = distritos[i % len(distritos)]
        nome = f"{rng.choice(NOMES)} {rng.choice(SOBRENOMES)}"
        if i % 23 == 0:                       # caixa alta, como vem digitado
            nome = nome.upper()
        if i % 31 == 0:                       # espaços sobrando
            nome = f"  {nome} "

        tem_coordenada = rng.random() < 0.74
        lat = round(clat + rng.uniform(-raio, raio), 6) if tem_coordenada else ""
        lon = round(clon + rng.uniform(-raio, raio), 6) if tem_coordenada else ""
        if tem_coordenada and i % 61 == 0:    # colunas trocadas
            lat, lon = lon, lat
        if tem_coordenada and i % 97 == 0:    # coordenada de outro estado
            lat = round(lat + 8.0, 6)

        linhas.append([
            f"{2026000 + i}",
            nome,
            rng.choice(RUAS),
            str(rng.randint(1, 900)) if tem_coordenada else "s/n",
            bairro,
            rng.choice(ESCOLAS).nome,
            rng.choice(TURNOS),
            rng.choices(CADEIRA, weights=PESOS_CADEIRA)[0],
            "sim" if rng.random() < 0.09 else "",
            lat, lon,
            rng.choice(OBSERVACOES),
        ])
        if i and i % 84 == 0:                 # aluno repetido em duas linhas
            linhas.append(list(linhas[-1]))
        if i and i % 110 == 0:                # linha em branco no meio
            linhas.append([""] * len(CABECALHO))
    return linhas


def _formatar_cabecalho(aba, linha_cabecalho: int, colunas: int):
    borda = Border(bottom=Side(style="thin", color=TINTA))
    for coluna in range(1, colunas + 1):
        celula = aba.cell(row=linha_cabecalho, column=coluna)
        celula.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        celula.fill = PatternFill("solid", fgColor=TINTA)
        celula.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
        celula.border = borda
    aba.row_dimensions[linha_cabecalho].height = 30
    aba.freeze_panes = aba.cell(row=linha_cabecalho + 1, column=1)


def _larguras(aba, larguras: dict):
    for coluna, largura in larguras.items():
        aba.column_dimensions[get_column_letter(coluna)].width = largura


def aba_alunos(planilha):
    aba = planilha.active
    aba.title = "Alunos 2026"

    aba["A1"] = "PREFEITURA MUNICIPAL DE RIBEIRÃO MODELO"
    aba["A2"] = ("Secretaria Municipal de Educação — Transporte Escolar 2026 "
                 "— relação de alunos transportados")
    for referencia, tamanho, negrito in (("A1", 14, True), ("A2", 10, False)):
        aba[referencia].font = Font(name="Arial", size=tamanho, bold=negrito,
                                    color=TINTA)
    aba.merge_cells(start_row=1, start_column=1, end_row=1,
                    end_column=len(CABECALHO))
    aba.merge_cells(start_row=2, start_column=1, end_row=2,
                    end_column=len(CABECALHO))

    linha_cabecalho = 4          # linha 3 fica em branco, como é de praxe
    for coluna, titulo in enumerate(CABECALHO, start=1):
        aba.cell(row=linha_cabecalho, column=coluna, value=titulo)
    _formatar_cabecalho(aba, linha_cabecalho, len(CABECALHO))

    linhas = _linhas_de_alunos()
    for i, linha in enumerate(linhas, start=linha_cabecalho + 1):
        for coluna, valor in enumerate(linha, start=1):
            celula = aba.cell(row=i, column=coluna, value=valor)
            celula.font = Font(name="Arial", size=10)
            if i % 2 == 0:
                celula.fill = PatternFill("solid", fgColor=CINZA)

    primeira, ultima = linha_cabecalho + 1, linha_cabecalho + len(linhas)
    rodape = ultima + 2
    aba.cell(row=rodape, column=1, value="TOTAL DE ALUNOS").font = Font(
        name="Arial", size=10, bold=True)
    # fórmula de verdade, como na planilha da secretaria — e que o importador
    # tem que reconhecer como rodapé, não como aluno
    aba.cell(row=rodape, column=2,
             value=f"=COUNTA(B{primeira}:B{ultima})").font = Font(
        name="Arial", size=10, bold=True)
    aba.cell(row=rodape + 1, column=1,
             value="Conferido por: ______________________  Data: ___/___/____"
             ).font = Font(name="Arial", size=9, italic=True)

    _larguras(aba, {1: 12, 2: 26, 3: 26, 4: 8, 5: 18, 6: 24, 7: 12, 8: 15,
                    9: 13, 10: 13, 11: 13, 12: 30})
    return aba


def aba_frota(planilha):
    """A frota que o município declara ter — o 'antes' da comparação."""
    aba = planilha.create_sheet("Frota atual")
    aba["A1"] = "Frota contratada / própria — Transporte Escolar 2026"
    aba["A1"].font = Font(name="Arial", size=12, bold=True, color=TINTA)

    cabecalho = ["Tipo de veículo", "Lugares", "Posições de cadeira de rodas",
                 "Quantidade", "Lugares totais"]
    for coluna, titulo in enumerate(cabecalho, start=1):
        aba.cell(row=3, column=coluna, value=titulo)
    _formatar_cabecalho(aba, 3, len(cabecalho))

    quantidades = {"ONIBUS31": 17, "MICRO20": 10, "VAN15A": 3}
    linha = 4
    for tipo in TIPOS_VEICULO:
        aba.cell(row=linha, column=1, value=tipo.nome)
        aba.cell(row=linha, column=2, value=tipo.capacidade)
        aba.cell(row=linha, column=3, value=tipo.posicoes_cadeirante)
        aba.cell(row=linha, column=4, value=quantidades.get(tipo.id, 0))
        aba.cell(row=linha, column=5, value=f"=B{linha}*D{linha}")
        for coluna in range(1, 6):
            aba.cell(row=linha, column=coluna).font = Font(name="Arial", size=10)
        linha += 1

    total = linha
    aba.cell(row=total, column=1, value="TOTAL").font = Font(
        name="Arial", size=10, bold=True)
    aba.cell(row=total, column=4, value=f"=SUM(D4:D{total - 1})").font = Font(
        name="Arial", size=10, bold=True)
    aba.cell(row=total, column=5, value=f"=SUM(E4:E{total - 1})").font = Font(
        name="Arial", size=10, bold=True)
    aba.cell(row=total + 2, column=1,
             value="Quilometragem declarada: 4.386 km/dia (relatório PNATE)"
             ).font = Font(name="Arial", size=9, italic=True)
    _larguras(aba, {1: 30, 2: 10, 3: 26, 4: 12, 5: 15})
    return aba


def main():
    planilha = Workbook()
    aba_alunos(planilha)
    aba_frota(planilha)
    planilha.save(SAIDA)
    print(f"Planilha de demonstração criada: {SAIDA}")
    return SAIDA


if __name__ == "__main__":
    main()
